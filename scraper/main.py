# ----------------------------- #
# Imports for Scraper
# ----------------------------- #
import os
import re
import time
import base64
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# ----------------------------- #
# Gmail API Imports
# ----------------------------- #
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build

# ----------------------------- #
# GLOBALS
# ----------------------------- #
BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # ✅ FIXED PATHS
CREDENTIALS_PATH = os.path.join(BASE_DIR, "credentials.json")
TOKEN_PATH = os.path.join(BASE_DIR, "token.json")

download_dir = os.path.join(os.getcwd(), "ppra_pdfs")
os.makedirs(download_dir, exist_ok=True)

options = webdriver.ChromeOptions()
options.add_argument("--headless=new")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument("--remote-debugging-port=9222")
options.add_argument("--window-size=1920,1080")
options.add_argument("--dns-prefetch-disable")
options.add_argument("--host-resolver-rules=MAP ppra.gov.pk 210.56.8.55")

prefs = {
    "download.default_directory": download_dir,
    "download.prompt_for_download": False,
    "plugins.always_open_pdf_externally": True,
    "profile.default_content_setting_values.automatic_downloads": 1
}
options.add_experimental_option("prefs", prefs)

service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=options)
driver.set_page_load_timeout(180)
wait = WebDriverWait(driver, 60)

url = "https://ppra.gov.pk/#/tenders/sectorwisetenders"

# ----------------------------- #
# SAFE NAVIGATION WITH RETRIES
# ----------------------------- #
def safe_get(url, retries=5, delay=15):
    for attempt in range(1, retries + 1):
        try:
            print(f"[Attempt {attempt}/{retries}] Navigating to {url}")
            driver.get(url)
            print("Page loaded successfully")
            return
        except Exception as e:
            print(f"Attempt {attempt} failed: {e}")
            if attempt == retries:
                print("All attempts failed. Exiting...")
                driver.quit()
                exit(1)
            print(f"Retrying in {delay} seconds...")
            time.sleep(delay)

safe_get(url)

tender_data = []
keyword_console_messages = []
detected_pdfs = []
matched_keywords_per_tender = {}

# ----------------------------- #
# SCRAPER FUNCTIONS
# ----------------------------- #
def wait_for_spinner_to_disappear(timeout=30):
    try:
        wait.until(EC.invisibility_of_element_located(
            (By.XPATH, "//div[contains(@class,'ngx-spinner-overlay')]")
        ))
    except TimeoutException:
        pass


def click_sector(sector_name="Info and Comm Tech"):
    found = False
    while not found:
        try:
            wait_for_spinner_to_disappear()
            sector_link = wait.until(EC.element_to_be_clickable(
                (By.XPATH, f"//a[contains(text(),'{sector_name}')]")
            ))
            driver.execute_script("arguments[0].click();", sector_link)
            found = True
        except TimeoutException:
            try:
                wait_for_spinner_to_disappear()
                next_btn = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, "//li[@class='page-item']/a[contains(text(),'Next')]")
                ))
                driver.execute_script("arguments[0].click();", next_btn)
                wait_for_spinner_to_disappear()
            except TimeoutException:
                driver.quit()
                exit()


def get_total_pages():
    try:
        total_pages_xpath = "//small[contains(text(),'Total Pages')]/strong"
        return int(wait.until(EC.presence_of_element_located((By.XPATH, total_pages_xpath))).text.strip())
    except:
        return 1


def scrape_page():
    rows_xpath = "//tr[contains(@class, 'ng-star-inserted')]"
    rows = driver.find_elements(By.XPATH, rows_xpath)
    for i in range(len(rows)):
        retry_count = 0
        while retry_count < 3:
            try:
                rows = driver.find_elements(By.XPATH, rows_xpath)
                row = rows[i]
                cols = row.find_elements(By.TAG_NAME, "td")
                if len(cols) >= 6:
                    sr_no = cols[0].text.strip()
                    tender_no = cols[1].text.strip().replace("View Invoice", "").strip()
                    tender_details = cols[2].text.strip()
                    adv_date = cols[4].text.strip()
                    close_date = cols[5].text.strip()
                    tender_data.append({
                        "Sr No": sr_no,
                        "Tender No": tender_no,
                        "Tender Details": tender_details,
                        "Advertisement Date": adv_date,
                        "Closing Date": close_date,
                    })
                    fourth_col = cols[3]
                    before_files = set(os.listdir(download_dir))
                    download_icons = fourth_col.find_elements(By.XPATH, ".//i[contains(@class,'fa-file-download')]")
                    for icon in download_icons:
                        try:
                            parent_a = icon.find_element(By.XPATH, "./parent::a")
                            href = parent_a.get_attribute("href")
                            if href == "javascript:void(0)":
                                continue
                            ext = ".pdf"
                            new_name = f"{tender_no}{ext}"
                            new_path = os.path.join(download_dir, new_name)
                            if os.path.exists(new_path):
                                continue
                            driver.execute_script("arguments[0].click();", icon)
                            timeout = 30
                            new_file = None
                            start_time = time.time()
                            while time.time() - start_time < timeout:
                                after_files = set(os.listdir(download_dir))
                                diff_files = after_files - before_files
                                if diff_files:
                                    new_file = diff_files.pop()
                                    break
                                time.sleep(0.5)
                            if new_file:
                                old_path = os.path.join(download_dir, new_file)
                                ext = os.path.splitext(new_file)[1]
                                new_path = os.path.join(download_dir, f"{tender_no}{ext}")
                                os.rename(old_path, new_path)
                        except:
                            pass
                break
            except:
                retry_count += 1
                time.sleep(1)
                if retry_count == 3:
                    break


def save_to_excel(excel_path):
    df = pd.DataFrame(tender_data)
    df.to_excel(excel_path, index=False)
    wb = load_workbook(excel_path)
    ws = wb.active

    tender_no_col = None
    details_col = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "Tender No":
            tender_no_col = idx
        if cell.value == "Tender Details":
            details_col = idx

    if tender_no_col and details_col:
        for row in ws.iter_rows(min_row=2):
            tender_no = row[tender_no_col - 1].value
            details_cell = row[details_col - 1]
            if tender_no in matched_keywords_per_tender:
                details_cell.font = Font(bold=True)

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_value = str(cell.value) if cell.value else ""
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 5, 100)

    for row in ws.iter_rows():
        max_lines = 1
        for cell in row:
            if cell.value:
                lines = str(cell.value).count("\n") + 1
                approx_line_length = len(str(cell.value)) // 50 + 1
                max_lines = max(max_lines, lines + approx_line_length)
        ws.row_dimensions[row[0].row].height = max_lines * 15

    wb.save(excel_path)


def detect_keyword(keyword_list):
    print(f"\nTenders with keywords : {keyword_list}\n")
    seen = set()
    for tender in tender_data:
        tender_no = tender.get("Tender No", "")
        details = tender.get("Tender Details", "")
        matched_keywords = [kw for kw in keyword_list if kw.lower() in details.lower()]
        if matched_keywords:
            key = tender_no
            if key not in seen:
                seen.add(key)
                matched_keywords_per_tender[tender_no] = matched_keywords

                formatted_details = details
                for kw in matched_keywords:
                    formatted_details = re.sub(
                        fr"(?i)({re.escape(kw)})",
                        r"<b>\1</b>",
                        formatted_details
                    )
                msg = (
                    f"<p><b>Tender {tender_no}</b> contains keywords "
                    f"<b>{', '.join(matched_keywords)}</b><br>"
                    f"------<br>{formatted_details}<br>------</p>"
                )
                print(msg)
                keyword_console_messages.append(msg)

                pdf_file = os.path.join(download_dir, f"{tender_no}.pdf")
                if os.path.exists(pdf_file):
                    detected_pdfs.append(pdf_file)

# ----------------------------- #
# EMAIL FUNCTIONS
# ----------------------------- #
SCOPES = ['https://www.googleapis.com/auth/gmail.send']

def gmail_authenticate():
    creds = None
    # ✅ Load token.json if exists
    if os.path.exists(TOKEN_PATH):
        creds = Credentials.from_authorized_user_file(TOKEN_PATH, SCOPES)

    # ✅ Refresh token or create a new one if missing
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            print("Creating a new token.json file...")
            flow = InstalledAppFlow.from_client_secrets_file(
                CREDENTIALS_PATH,
                SCOPES
            )
            creds = flow.run_local_server(port=0)
        with open(TOKEN_PATH, 'w') as token:
            token.write(creds.to_json())
    return build('gmail', 'v1', credentials=creds)


def create_message_with_attachments(sender, to, subject, body_text, files=[]):
    message = MIMEMultipart()
    message['to'] = to
    message['from'] = sender
    message['subject'] = subject
    message.attach(MIMEText(body_text, 'html'))

    for file in files:
        with open(file, 'rb') as f:
            mime_base = MIMEBase('application', 'octet-stream')
            mime_base.set_payload(f.read())
        encoders.encode_base64(mime_base)
        mime_base.add_header(
            'Content-Disposition',
            f'attachment; filename="{os.path.basename(file)}"'
        )
        message.attach(mime_base)

    raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode()
    return {'raw': raw_message}


def send_email(sender_email, recipient_email, subject, body_text, excel_path, pdf_files):
    service = gmail_authenticate()
    files_to_send = [excel_path] + pdf_files
    message = create_message_with_attachments(
        sender_email,
        recipient_email,
        subject,
        body_text,
        files_to_send
    )
    service.users().messages().send(userId="me", body=message).execute()
    print("Email sent successfully.")

# ----------------------------- #
# MAIN FUNCTION
# ----------------------------- #
def main():
    wait_for_spinner_to_disappear()
    click_sector("Info and Comm Tech")
    wait_for_spinner_to_disappear()
    wait.until(EC.presence_of_element_located((By.XPATH, "//tr[contains(@class, 'ng-star-inserted')]")))

    total_pages = get_total_pages()
    for page in range(1, total_pages + 1):
        scrape_page()
        if page < total_pages:
            try:
                wait_for_spinner_to_disappear()
                next_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[contains(text(),'Next')]")))
                driver.execute_script("arguments[0].click();", next_btn)
                time.sleep(1)
            except TimeoutException:
                break
    driver.quit()

    excel_path = os.path.join(BASE_DIR, "ppra_info_comm_tech.xlsx")
    save_to_excel(excel_path)

    print(f"\nData saved to {excel_path}")
    print(f"All PDFs in {download_dir}")

    keywords = ["Bank", "University"]
    detect_keyword(keywords)

    sender_email = os.getenv("GMAIL_SENDER")
    recipient_email = os.getenv("GMAIL_RECIPIENT")

    if keyword_console_messages:
        email_body = "".join(keyword_console_messages)
        subject = f"PPRA Tender Report - {len(keyword_console_messages)} matches found"
    else:
        email_body = "<p><b>No keywords detected.</b></p>"
        subject = "PPRA Tender Report - No matches"

    send_email(sender_email, recipient_email, subject, email_body, excel_path, detected_pdfs)


if __name__ == "__main__":
    main()
